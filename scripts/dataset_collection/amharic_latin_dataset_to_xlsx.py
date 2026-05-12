"""
Скачивает амхарский новостной датасет с HuggingFace,
берёт 2500 строк (от 150 символов), транслитерирует геэз → латиницу.

Датасет: rasyosef/amharic-news-category-classification
Колонка: headline

Запуск:
    py scripts/dataset_collection/amharic_latin_dataset_to_xlsx.py

Результат: data/amharic_latin_dataset.xlsx
  Лист 1 — Amharic (Original):  2500 строк на геэз
  Лист 2 — Amharic (Latin):     2500 строк на латинице
"""

from pathlib import Path

from datasets import load_dataset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
MAX_CHARS = 300
MIN_CHARS = 150
N = 2500


def _build_translit_map():
    # BGN/PCGN 1967 Romanization of Amharic
    # Each tuple: (geez_start_codepoint, 1st, 2nd, 3rd, 4th, 5th, 6th, 7th[, wa_combo])
    # 6th order = consonant+i (simplified without i per BGN note 1)
    _ROWS = [
        (0x1200, "hā",  "hu",  "hī",  "ha",  "hē",  "h",  "ho"),
        (0x1208, "le",  "lu",  "lī",  "la",  "lē",  "l",  "lo",  "lwa"),
        (0x1210, "hā",  "hu",  "hī",  "ha",  "hē",  "h",  "ho"),
        (0x1218, "me",  "mu",  "mī",  "ma",  "mē",  "m",  "mo",  "mwa"),
        (0x1220, "se",  "su",  "sī",  "sa",  "sē",  "s",  "so"),
        (0x1228, "re",  "ru",  "rī",  "ra",  "rē",  "r",  "ro",  "rwa"),
        (0x1230, "se",  "su",  "sī",  "sa",  "sē",  "s",  "so",  "swa"),
        (0x1238, "she", "shu", "shī", "sha", "shē", "sh", "sho", "shwa"),
        (0x1240, "k'e", "k'u", "k'ī", "k'a", "k'ē", "k'", "k'o", "k'wa"),
        (0x1250, "k'e", "k'u", "k'ī", "k'a", "k'ē", "k'", "k'o", "k'wa"),
        (0x1260, "be",  "bu",  "bī",  "ba",  "bē",  "b",  "bo",  "bwa"),
        (0x1268, "ve",  "vu",  "vī",  "va",  "vē",  "v",  "vo"),
        (0x1270, "te",  "tu",  "tī",  "ta",  "tē",  "t",  "to",  "twa"),
        (0x1278, "che", "chu", "chī", "cha", "chē", "ch", "cho", "chwa"),
        (0x1280, "hā",  "hu",  "hī",  "ha",  "hē",  "h",  "ho",  "hwa"),
        (0x1290, "ne",  "nu",  "nī",  "na",  "nē",  "n",  "no",  "nwa"),
        (0x1298, "nye", "nyu", "nyī", "nya", "nyē", "ny", "nyo", "nywa"),
        (0x12A0, "ā",   "u",   "ī",   "a",   "ē",   "i",  "o"),
        (0x12A8, "ke",  "ku",  "kī",  "ka",  "kē",  "k",  "ko",  "kwa"),
        (0x12B8, "he",  "hu",  "hī",  "ha",  "hē",  "h",  "ho",  "hwa"),
        (0x12C0, "we",  "wu",  "wī",  "wa",  "wē",  "w",  "wo"),
        (0x12C8, "ze",  "zu",  "zī",  "za",  "zē",  "z",  "zo",  "zwa"),
        (0x12D0, "zhe", "zhu", "zhī", "zha", "zhē", "zh", "zho", "zhwa"),
        (0x12D8, "ye",  "yu",  "yī",  "ya",  "yē",  "y",  "yo"),
        (0x12E0, "de",  "du",  "dī",  "da",  "dē",  "d",  "do",  "dwa"),
        (0x12E8, "je",  "ju",  "jī",  "ja",  "jē",  "j",  "jo",  "jwa"),
        (0x12F0, "ge",  "gu",  "gī",  "ga",  "gē",  "g",  "go",  "gwa"),
        (0x1300, "t'e", "t'u", "t'ī", "t'a", "t'ē", "t'", "t'o", "t'wa"),
        (0x1308, "ch'e","ch'u","ch'ī","ch'a","ch'ē","ch'","ch'o","ch'wa"),
        (0x1310, "p'e", "p'u", "p'ī", "p'a", "p'ē", "p'", "p'o"),
        (0x1318, "ts'e","ts'u","ts'ī","ts'a","ts'ē","ts'","ts'o", "ts'wa"),
        (0x1328, "fe",  "fu",  "fī",  "fa",  "fē",  "f",  "fo",  "fwa"),
        (0x1330, "pe",  "pu",  "pī",  "pa",  "pē",  "p",  "po"),
    ]
    _WA = {
        0x1240: (0x1248, "k'wa"),  0x1250: (0x1258, "k'wa"),
        0x1280: (0x1288, "hwa"),   0x12A8: (0x12B0, "kwa"),
        0x12F0: (0x12F8, "gwa"),
    }
    m = {}
    for row in _ROWS:
        base = row[0]
        forms = row[1:]
        for vi, latin in enumerate(forms):
            m[chr(base + vi)] = latin
        if base in _WA:
            wa_base, wa_val = _WA[base]
            for vi in range(7):
                m[chr(wa_base + vi)] = wa_val
    m[chr(0x1361)] = " "
    m[chr(0x1362)] = ". "
    m[chr(0x1363)] = ", "
    m[chr(0x1364)] = "; "
    m[chr(0x1365)] = ": "
    m[chr(0x1366)] = ": "
    m[chr(0x1367)] = "? "
    m[chr(0x1368)] = " "
    return m


TRANSLIT_MAP = _build_translit_map()


def transliterate(text: str) -> str:
    missed = set()
    result = []
    for ch in text:
        t = TRANSLIT_MAP.get(ch)
        if t is not None:
            result.append(t)
        else:
            if '\u1200' <= ch <= '\u137f':
                missed.add(f"U+{ord(ch):04X}")
            result.append(ch)
    if missed:
        print(f"  WARNING: непереведённые геэз-символы: {sorted(missed)}")
    return "".join(result)


def truncate_text(text: str) -> str:
    text = text.strip()
    if len(text) <= MAX_CHARS:
        return text
    truncated = text[:MAX_CHARS]
    last_space = truncated.rfind(" ")
    if last_space > MAX_CHARS // 2:
        return truncated[:last_space].strip()
    return truncated.strip()


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="006400")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
ROW_FONT = Font(name="Arial", size=10)
ROW_ALIGN = Alignment(wrap_text=True, vertical="top")


def style_sheet(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def write_rows(ws, rows):
    for r_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 30
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = ROW_FONT
            cell.alignment = ROW_ALIGN


print("Загружаем датасет rasyosef/amharic-news-category-classification ...")
ds = load_dataset("rasyosef/amharic-news-category-classification")

split = list(ds.keys())[0]
data = ds[split]

col_names = data.column_names
print(f"Колонки: {col_names}")

text_col = "headline"
if text_col not in col_names:
    text_col = next(
        (c for c in col_names if c.lower() in {"headline", "text", "title"}),
        col_names[0],
    )
print(f"Текстовая колонка: '{text_col}'")

extra_cols = [c for c in col_names if c != text_col]

rows_original = []
rows_latin = []

print(f"Собираем {N} строк (минимум {MIN_CHARS} символов)...")
idx = 1
for example in data:
    if len(rows_original) >= N:
        break
    text = str(example[text_col]).strip()
    if not text or len(text) < MIN_CHARS:
        continue
    text = truncate_text(text)
    row_meta = [example[c] for c in extra_cols]
    rows_original.append([idx, text] + row_meta)
    rows_latin.append([idx, transliterate(text)] + row_meta)
    if idx % 500 == 0:
        print(f"  {idx}/{N}...")
    idx += 1

if len(rows_original) < N:
    print(f"  Собрано {len(rows_original)}/{N} (>={MIN_CHARS} символов). Добавляем более короткие...")
    for example in data:
        if len(rows_original) >= N:
            break
        text = str(example[text_col]).strip()
        if not text or len(text) >= MIN_CHARS:
            continue
        text = truncate_text(text)
        row_meta = [example[c] for c in extra_cols]
        rows_original.append([idx, text] + row_meta)
        rows_latin.append([idx, transliterate(text)] + row_meta)
        if idx % 500 == 0:
            print(f"  {idx}/{N}...")
        idx += 1

print(f"Строк собрано: {len(rows_original)}")

print("\nПримеры транслитерации:")
for i in range(min(3, len(rows_original))):
    orig = rows_original[i][1][:80]
    lat = rows_latin[i][1][:80]
    print(f"  {orig}")
    print(f"  → {lat}\n")

wb = Workbook()

col_w = {
    "A": 7,
    "B": 90,
    **{get_column_letter(i + 3): 15 for i in range(len(extra_cols))},
}

ws1 = wb.active
ws1.title = "Amharic (Original)"
ws1.append(["#", text_col] + extra_cols)
write_rows(ws1, rows_original)
style_sheet(ws1, col_w)

ws2 = wb.create_sheet("Amharic (Latin)")
ws2.append(["#", f"{text_col}_latin"] + extra_cols)
write_rows(ws2, rows_latin)
style_sheet(ws2, col_w)

out_path = PROJECT_ROOT / "data" / "amharic_latin_dataset.xlsx"
tmp_path = out_path.with_suffix(".tmp.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(tmp_path)
if out_path.exists():
    out_path.unlink()
tmp_path.rename(out_path)
print(f"Готово! Файл сохранён: {out_path}")
print(f"  Лист 1 «Amharic (Original)»  — {len(rows_original)} строк")
print(f"  Лист 2 «Amharic (Latin)»     — {len(rows_latin)} строк")
