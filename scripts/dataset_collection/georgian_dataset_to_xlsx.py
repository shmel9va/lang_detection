"""
Скачивает грузинский датасет с HuggingFace, берёт 2500 строк на грузинском,
транслитерирует на латиницу по Национальной системе романизации Грузии (2002).

Запуск:
    py scripts/dataset_collection/georgian_dataset_to_xlsx.py

Результат: data/georgian_dataset.xlsx
  Лист 1 — Georgian (Original):       2500 строк на мхедрули
  Лист 2 — Georgian (Transliterated): 2500 строк на латинице
"""

import os
from pathlib import Path
from datasets import load_dataset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]  # .../lang_detection
TEXT_COLS    = {'text', 'sentence', 'review', 'comment'}

# ── Транслитерация грузинского → латиница ────────────────────────────────────
# Национальная система романизации Грузии (2002), используется на паспортах
# и дорожных знаках. Источник: Указ № 575, 2002.
TRANSLIT_MAP = {
    'ა': 'a',  'ბ': 'b',  'გ': 'g',  'დ': 'd',  'ე': 'e',
    'ვ': 'v',  'ზ': 'z',  'თ': 't',  'ი': 'i',  'კ': 'k',
    'ლ': 'l',  'მ': 'm',  'ნ': 'n',  'ო': 'o',  'პ': 'p',
    'ჟ': 'zh', 'რ': 'r',  'ს': 's',  'ტ': 't',  'უ': 'u',
    'ფ': 'p',  'ქ': 'k',  'ღ': 'gh', 'ყ': 'q',  'შ': 'sh',
    'ჩ': 'ch', 'ც': 'ts', 'ძ': 'dz', 'წ': 'ts', 'ჭ': 'ch',
    'ხ': 'kh', 'ჯ': 'j',  'ჰ': 'h',
}

def transliterate(text: str) -> str:
    return ''.join(TRANSLIT_MAP.get(ch, ch) for ch in text)


# ── Стиль ─────────────────────────────────────────────────────────────────────
HEADER_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL  = PatternFill('solid', start_color='1B5E20')  # тёмно-зелёный
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
ROW_FONT     = Font(name='Arial', size=10)
ROW_ALIGN    = Alignment(wrap_text=True, vertical='top')


def style_sheet(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def write_rows(ws, rows):
    for r_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 30
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = ROW_FONT
            cell.alignment = ROW_ALIGN


# ── Загрузка датасета ─────────────────────────────────────────────────────────
print("Загружаем датасет Arseniy-Sandalov/Georgian-Sentiment-Analysis (streaming) ...")
ds = load_dataset("Arseniy-Sandalov/Georgian-Sentiment-Analysis", streaming=True)

split = list(ds.keys())[0]
data  = ds[split]

col_names = data.column_names
print(f"Колонки: {col_names}")

# Находим текстовую колонку
text_col = next(
    (c for c in col_names if c.lower() in TEXT_COLS),
    col_names[0],  # берём первую, если стандартное имя не найдено
)
print(f"Текстовая колонка: '{text_col}'")

extra_cols = [c for c in col_names if c != text_col]

N = 2500
rows_original       = []
rows_transliterated = []

idx = 1
for example in data:
    if len(rows_original) >= N:
        break
    text = str(example[text_col]).strip()
    if not text:
        continue
    row_meta = [example[c] for c in extra_cols]
    rows_original.append([idx, text] + row_meta)
    rows_transliterated.append([idx, transliterate(text)] + row_meta)
    idx += 1

print(f"Строк собрано: {len(rows_original)}")

# ── Создаём Workbook ──────────────────────────────────────────────────────────
wb = Workbook()

header_base = ['#', text_col] + extra_cols

ws1 = wb.active
ws1.title = "Georgian (Original)"
ws1.append(header_base)
write_rows(ws1, rows_original)
col_w = {'A': 6, 'B': 90, **{get_column_letter(i + 3): 15 for i in range(len(extra_cols))}}
style_sheet(ws1, col_w)

ws2 = wb.create_sheet("Georgian (Transliterated Latin)")
ws2.append(['#', f'{text_col}_latin'] + extra_cols)
write_rows(ws2, rows_transliterated)
style_sheet(ws2, col_w)

# ── Сохраняем ─────────────────────────────────────────────────────────────────
out_path = PROJECT_ROOT / "data" / "georgian_dataset.xlsx"
tmp_path = out_path.with_suffix(".tmp.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(tmp_path)
if out_path.exists():
    out_path.unlink()
tmp_path.rename(out_path)
print(f"Готово! Файл сохранён: {out_path}")
print(f"  Лист 1 «Georgian (Original)»           — {len(rows_original)} строк")
print(f"  Лист 2 «Georgian (Transliterated Latin)» — {len(rows_transliterated)} строк")
