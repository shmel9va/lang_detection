"""
Скачивает непальский датасет с HuggingFace, берёт 2500 строк,
транслитерирует деванагари → латиницу через пакет nepali-to-roman.

Установка зависимости (один раз):
    pip install nepali-to-roman

Запуск:
    py scripts/dataset_collection/nepali_dataset_to_xlsx.py

Результат: data/nepali_dataset.xlsx
  Лист 1 — Nepali (Original):         2500 строк на деванагари
  Лист 2 — Nepali (Transliterated):   2500 строк на латинице
"""

from pathlib import Path
from datasets import load_dataset
import ntr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
TEXT_COLS    = {'text', 'sentence', 'content', 'body'}


def transliterate(text: str) -> str:
    try:
        return ntr.nep_to_rom(text)
    except Exception:
        return text


# ── Стиль ─────────────────────────────────────────────────────────────────────
HEADER_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL  = PatternFill('solid', start_color='6A1B9A')  # фиолетовый
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
ROW_FONT     = Font(name='Arial', size=10)
ROW_ALIGN    = Alignment(wrap_text=True, vertical='top')


def style_sheet(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def write_rows(ws, rows):
    for r_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 30
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = ROW_FONT
            cell.alignment = ROW_ALIGN


# ── Загрузка датасета ─────────────────────────────────────────────────────────
print("Загружаем датасет Sakonii/nepalitext-language-model-dataset (streaming) ...")
ds = load_dataset("Sakonii/nepalitext-language-model-dataset", streaming=True)

split = list(ds.keys())[0]
data  = ds[split]

col_names = data.column_names
print(f"Колонки: {col_names}")

text_col = next(
    (c for c in col_names if c.lower() in TEXT_COLS),
    col_names[0],
)
print(f"Текстовая колонка: '{text_col}'")

extra_cols = [c for c in col_names if c != text_col]

N = 2500
rows_original       = []
rows_transliterated = []

print(f"Обрабатываем {N} строк (транслитерация может занять ~30 сек)...")
idx = 1
for example in data:
    if len(rows_original) >= N:
        break
    text = str(example[text_col]).strip()
    if not text:
        continue
    row_meta = [example[c] for c in extra_cols]
    rows_original.append([idx, text] + row_meta)
    rows_transliterated.append([idx, transliterate(text)] + row_meta)
    if idx % 500 == 0:
        print(f"  {idx}/{N}...")
    idx += 1

print(f"Строк собрано: {len(rows_original)}")

# ── Создаём Workbook ──────────────────────────────────────────────────────────
wb = Workbook()

col_w = {'A': 6, 'B': 90, **{get_column_letter(i + 3): 15 for i in range(len(extra_cols))}}

ws1 = wb.active
ws1.title = "Nepali (Original)"
ws1.append(['#', text_col] + extra_cols)
write_rows(ws1, rows_original)
style_sheet(ws1, col_w)

ws2 = wb.create_sheet("Nepali (Transliterated Latin)")
ws2.append(['#', f'{text_col}_latin'] + extra_cols)
write_rows(ws2, rows_transliterated)
style_sheet(ws2, col_w)

# ── Сохраняем ─────────────────────────────────────────────────────────────────
out_path = PROJECT_ROOT / "data" / "nepali_dataset.xlsx"
tmp_path = out_path.with_suffix(".tmp.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(tmp_path)
if out_path.exists():
    out_path.unlink()
tmp_path.rename(out_path)
print(f"Готово! Файл сохранён: {out_path}")
print(f"  Лист 1 «Nepali (Original)»           — {len(rows_original)} строк")
print(f"  Лист 2 «Nepali (Transliterated Latin)» — {len(rows_transliterated)} строк")
