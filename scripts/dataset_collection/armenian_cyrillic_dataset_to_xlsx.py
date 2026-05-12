"""
Скачивает армянский парафразный датасет с HuggingFace,
берёт 2500 строк, транслитерирует армянский алфавит → кириллицу.

Датасет: Karavet/ARPA-Armenian-Paraphrase-Corpus

Запуск:
    py scripts/dataset_collection/armenian_cyrillic_dataset_to_xlsx.py

Результат: data/armenian_cyrillic_dataset.xlsx
  Лист 1 — Armenian (Original):  2500 строк на армянском алфавите
  Лист 2 — Armenian (Cyrillic):  2500 строк на кириллице
"""

from pathlib import Path

from datasets import load_dataset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
TEXT_COLS = {
    "text", "sentence", "source", "target", "armenian",
    "content", "sentence_1", "sentence_2", "input", "query",
}
MAX_CHARS = 300
N = 2500

# ── Транслитерация армянский алфавит → кириллица (фонетическая) ──────────
TRANSLIT_MAP = {
    "ա": "а", "բ": "б", "գ": "г", "դ": "д", "ե": "е", "զ": "з",
    "է": "э", "ը": "ы", "թ": "т", "ժ": "ж", "ի": "и", "լ": "л",
    "խ": "х", "ծ": "ц", "կ": "к", "հ": "h", "ձ": "дз", "ղ": "гх",
    "ճ": "ч", "մ": "м", "յ": "й", "ն": "н", "շ": "ш", "ո": "о",
    "չ": "ч", "պ": "п", "ջ": "дж", "ռ": "р", "ս": "с", "վ": "в",
    "տ": "т", "ր": "р", "ւ": "в", "ց": "ц", "փ": "п", "ք": "к",
    "օ": "о", "ֆ": "ф", "և": "ев",
    "Ա": "А", "Բ": "Б", "Գ": "Г", "Դ": "Д", "Ե": "Е", "Զ": "З",
    "Է": "Э", "Ը": "Ы", "Թ": "Т", "Ժ": "Ж", "Ի": "И", "Լ": "Л",
    "Խ": "Х", "Ծ": "Ц", "Կ": "К", "Հ": "H", "Ձ": "Дз", "Ղ": "Гх",
    "Ճ": "Ч", "Մ": "М", "Յ": "Й", "Ն": "Н", "Շ": "Ш", "Ո": "О",
    "Չ": "Ч", "Պ": "П", "Ջ": "Дж", "Ռ": "Р", "Ս": "С", "Վ": "В",
    "Տ": "Т", "Ր": "Р", "Ց": "Ц", "Փ": "П", "Ք": "К", "Օ": "О",
    "Ֆ": "Ф",
}


def transliterate(text: str) -> str:
    result = []
    i = 0
    while i < len(text):
        if i + 1 < len(text):
            pair = text[i : i + 2]
            if pair == "ու":
                result.append("у")
                i += 2
                continue
            if pair == "Ու":
                result.append("У")
                i += 2
                continue
            if pair == "Եվ":
                result.append("Ев")
                i += 2
                continue
        result.append(TRANSLIT_MAP.get(text[i], text[i]))
        i += 1
    return "".join(result)


def truncate_text(text: str, max_chars: int = MAX_CHARS) -> str:
    text = text.strip()
    if len(text) <= max_chars:
        return text
    truncated = text[:max_chars]
    last_space = truncated.rfind(" ")
    if last_space > max_chars // 2:
        return truncated[:last_space].strip()
    return truncated.strip()


# ── Стиль ─────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="8B0000")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
ROW_FONT = Font(name="Arial", size=10)
ROW_ALIGN = Alignment(wrap_text=True, vertical="top")


def style_sheet(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def write_rows(ws, rows):
    for r_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 30
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = ROW_FONT
            cell.alignment = ROW_ALIGN


# ── Загрузка датасета ─────────────────────────────────────────────────────
print("Загружаем датасет Karavet/ARPA-Armenian-Paraphrase-Corpus ...")
ds = load_dataset("Karavet/ARPA-Armenian-Paraphrase-Corpus")

split = list(ds.keys())[0]
data = ds[split]

col_names = data.column_names
print(f"Колонки: {col_names}")

text_col = next(
    (c for c in col_names if c.lower() in TEXT_COLS),
    col_names[0],
)
print(f"Текстовая колонка: '{text_col}'")

extra_cols = [c for c in col_names if c != text_col]

rows_original = []
rows_cyrillic = []

print(f"Обрабатываем {N} строк (обрезка до {MAX_CHARS} символов + транслитерация)...")
idx = 1
for example in data:
    if len(rows_original) >= N:
        break
    text = str(example[text_col]).strip()
    if not text:
        continue
    text = truncate_text(text)
    row_meta = [example[c] for c in extra_cols]
    rows_original.append([idx, text] + row_meta)
    rows_cyrillic.append([idx, transliterate(text)] + row_meta)
    if idx % 500 == 0:
        print(f"  {idx}/{N}...")
    idx += 1

print(f"Строк собрано: {len(rows_original)}")

print("\nПримеры транслитерации:")
for i in range(min(3, len(rows_original))):
    orig = rows_original[i][1][:80]
    cyr = rows_cyrillic[i][1][:80]
    print(f"  {orig}")
    print(f"  → {cyr}\n")

# ── Создаём Workbook ──────────────────────────────────────────────────────
wb = Workbook()

col_w = {
    "A": 7,
    "B": 90,
    **{get_column_letter(i + 3): 15 for i in range(len(extra_cols))},
}

ws1 = wb.active
ws1.title = "Armenian (Original)"
ws1.append(["#", text_col] + extra_cols)
write_rows(ws1, rows_original)
style_sheet(ws1, col_w)

ws2 = wb.create_sheet("Armenian (Cyrillic)")
ws2.append(["#", f"{text_col}_cyrillic"] + extra_cols)
write_rows(ws2, rows_cyrillic)
style_sheet(ws2, col_w)

# ── Сохраняем ─────────────────────────────────────────────────────────────
out_path = PROJECT_ROOT / "data" / "armenian_cyrillic_dataset.xlsx"
tmp_path = out_path.with_suffix(".tmp.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(tmp_path)
if out_path.exists():
    out_path.unlink()
tmp_path.rename(out_path)
print(f"Готово! Файл сохранён: {out_path}")
print(f"  Лист 1 «Armenian (Original)»  — {len(rows_original)} строк")
print(f"  Лист 2 «Armenian (Cyrillic)»  — {len(rows_cyrillic)} строк")
