import os
from pathlib import Path
from datasets import load_dataset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]  # .../lang_detection
TEXT_COLS    = {'sentence', 'text', 'armenian_translation'}

# ── Транслитерация армянского → латиница ──────────────────────────────────────
TRANSLIT_MAP = {
    'ա': 'a',  'բ': 'b',  'գ': 'g',  'դ': 'd',  'ե': 'ye', 'զ': 'z',
    'է': 'e',  'ը': 'ë',  'թ': "t'", 'ժ': 'zh', 'ի': 'i',  'լ': 'l',
    'խ': 'kh', 'ծ': 'ts', 'կ': 'k',  'հ': 'h',  'ձ': 'dz', 'ղ': 'gh',
    'չ': 'ch', 'մ': 'm',  'յ': 'y',  'ն': 'n',  'շ': 'sh', 'ո': 'vo',
    'չ': 'ch', 'պ': 'p',  'ջ': 'j',  'ռ': 'rr', 'ս': 's',  'վ': 'v',
    'տ': 't',  'ր': 'r',  'ց': 'ts', 'ու': 'u', 'փ': "p'", 'ք': "k'",
    'օ': 'o',  'ֆ': 'f',  'ե': 'e',  'և': 'ev', 'ω': 'o',
    'Ա': 'A',  'Բ': 'B',  'Գ': 'G',  'Դ': 'D',  'Ե': 'Ye', 'Զ': 'Z',
    'Է': 'E',  'Ը': 'Ë',  'Թ': "T'", 'Ժ': 'Zh', 'Ի': 'I',  'Լ': 'L',
    'Խ': 'Kh', 'Ծ': 'Ts', 'Կ': 'K',  'Հ': 'H',  'Ձ': 'Dz', 'Ղ': 'Gh',
    'Չ': 'Ch', 'Մ': 'M',  'Յ': 'Y',  'Ն': 'N',  'Շ': 'Sh', 'Ո': 'Vo',
    'Պ': 'P',  'Ջ': 'J',  'Ռ': 'Rr', 'Ս': 'S',  'Վ': 'V',  'Տ': 'T',
    'Ր': 'R',  'Ց': 'Ts', 'Փ': "P'", 'Ք': "K'", 'Օ': 'O',  'Ֆ': 'F',
    'Եվ': 'Ev',
}

def transliterate(text: str) -> str:
    result = []
    i = 0
    while i < len(text):
        # двухсимвольные диграфы — ու и Եվ
        pair = text[i:i+2]
        if pair == 'ու':
            result.append('u')
            i += 2
        elif pair == 'Եվ':
            result.append('Ev')
            i += 2
        else:
            result.append(TRANSLIT_MAP.get(text[i], text[i]))
            i += 1
    return ''.join(result)


# ── Стиль шапки ───────────────────────────────────────────────────────────────
HEADER_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL   = PatternFill('solid', start_color='1F4E79')
HEADER_ALIGN  = Alignment(horizontal='center', vertical='center')
ROW_FONT      = Font(name='Arial', size=10)
ROW_ALIGN     = Alignment(wrap_text=True, vertical='top')


def style_sheet(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def write_rows(ws, rows):
    for r_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 30
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = ROW_FONT
            cell.alignment = ROW_ALIGN


# ── Загрузка датасета ─────────────────────────────────────────────────────────
print("Загружаем датасет EdUarD0110/Armenian_sentences_100000 ...")
ds = load_dataset("EdUarD0110/Armenian_sentences_100000")

# Определяем сплит (обычно 'train')
split = list(ds.keys())[0]
data  = ds[split]

# Берём первые 5000 строк: 2500 для листа 1, следующие 2500 для листа 2
N = 2500
rows_original    = []
rows_transliterated = []

col_names = data.column_names
print(f"Колонки: {col_names}")

for i, example in enumerate(data):
    if i >= N * 2:
        break

    # Берём текст из первой текстовой колонки
    text = str(example.get('sentence') or example.get('text') or example.get('armenian_translation') or list(example.values())[0])

    row_base = [i + 1, text]
    for key in col_names:
        if key not in TEXT_COLS:
            row_base.append(example[key])

    if i < N:
        rows_original.append(row_base)
    else:
        rows_transliterated.append([i - N + 1, transliterate(text)] +
                                    [example[k] for k in col_names
                                     if k not in TEXT_COLS])

# ── Создаём Workbook ──────────────────────────────────────────────────────────
wb = Workbook()

# Лист 1 — оригинальный армянский алфавит
ws1 = wb.active
ws1.title = "Armenian (Original)"
extra_cols = [k for k in col_names if k not in TEXT_COLS]
header1 = ['#', 'sentence'] + extra_cols
ws1.append(header1)
write_rows(ws1, rows_original)
style_sheet(ws1, {'A': 7, 'B': 90, **{get_column_letter(i+3): 15
                                        for i in range(len(extra_cols))}})

# Лист 2 — транслитерированный (латиница)
ws2 = wb.create_sheet("Armenian (Transliterated Latin)")
header2 = ['#', 'sentence_latin'] + extra_cols
ws2.append(header2)
write_rows(ws2, rows_transliterated)
style_sheet(ws2, {'A': 7, 'B': 90, **{get_column_letter(i+3): 15
                                        for i in range(len(extra_cols))}})

# ── Сохраняем ─────────────────────────────────────────────────────────────────
out_path = PROJECT_ROOT / "data" / "armenian_dataset.xlsx"
tmp_path = out_path.with_suffix(".tmp.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(tmp_path)
if out_path.exists():
    out_path.unlink()
tmp_path.rename(out_path)
print(f"Готово! Файл сохранён: {out_path}")
print(f"  Лист 1 «Armenian (Original)»           — {len(rows_original)} строк")
print(f"  Лист 2 «Armenian (Transliterated Latin)» — {len(rows_transliterated)} строк")