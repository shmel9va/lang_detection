"""
Скачивает русский датасет с HuggingFace,
берёт 2500 строк, транслитерирует кириллицу → латиницу.

Датасет: MonoHime/ru_sentiment_dataset

Запуск:
    py scripts/dataset_collection/russian_latin_dataset_to_xlsx.py

Результат: data/russian_latin_dataset.xlsx
  Лист 1 — Russian (Original):  2500 строк на кириллице
  Лист 2 — Russian (Latin):     2500 строк на латинице
"""

from pathlib import Path

from datasets import load_dataset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
TEXT_COLS = {
    "text", "sentence", "content", "review", "comment",
    "message", "description", "body", "input",
}
MAX_CHARS = 300
N = 2500

# ── Транслитерация русский (кириллица → латиница) ────────────────────────
TRANSLIT_MAP = {
    "а": "a",  "б": "b",  "в": "v",  "г": "g",  "д": "d",
    "е": "e",  "ё": "yo", "ж": "zh", "з": "z",  "и": "i",
    "й": "y",  "к": "k",  "л": "l",  "м": "m",  "н": "n",
    "о": "o",  "п": "p",  "р": "r",  "с": "s",  "т": "t",
    "у": "u",  "ф": "f",  "х": "kh", "ц": "ts", "ч": "ch",
    "ш": "sh", "щ": "shch", "ъ": "",  "ы": "y",  "ь": "",
    "э": "e",  "ю": "yu", "я": "ya",
    "А": "A",  "Б": "B",  "В": "V",  "Г": "G",  "Д": "D",
    "Е": "E",  "Ё": "Yo", "Ж": "Zh", "З": "Z",  "И": "I",
    "Й": "Y",  "К": "K",  "Л": "L",  "М": "M",  "Н": "N",
    "О": "O",  "П": "P",  "Р": "R",  "С": "S",  "Т": "T",
    "У": "U",  "Ф": "F",  "Х": "Kh", "Ц": "Ts", "Ч": "Ch",
    "Ш": "Sh", "Щ": "Shch", "Ъ": "",  "Ы": "Y",  "Ь": "",
    "Э": "E",  "Ю": "Yu", "Я": "Ya",
}


def transliterate(text: str) -> str:
    return "".join(TRANSLIT_MAP.get(ch, ch) for ch in text)


def truncate_text(text: str, max_chars: int = MAX_CHARS) -> str:
    text = text.strip()
    if len(text) <= max_chars:
        return text
    truncated = text[:max_chars]
    last_space = truncated.rfind(" ")
    if last_space > max_chars // 2:
        return truncated[:last_space].strip()
    return truncated.strip()


# ── Стиль ─────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1A237E")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
ROW_FONT = Font(name="Arial", size=10)
ROW_ALIGN = Alignment(wrap_text=True, vertical="top")


def style_sheet(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def write_rows(ws, rows):
    for r_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 30
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = ROW_FONT
            cell.alignment = ROW_ALIGN


# ── Загрузка датасета ─────────────────────────────────────────────────────
print("Загружаем датасет MonoHime/ru_sentiment_dataset ...")
ds = load_dataset("MonoHime/ru_sentiment_dataset")

split = list(ds.keys())[0]
data = ds[split]

col_names = data.column_names
print(f"Колонки: {col_names}")

text_col = next(
    (c for c in col_names if c.lower() in TEXT_COLS),
    col_names[0],
)
print(f"Текстовая колонка: '{text_col}'")

extra_cols = [c for c in col_names if c != text_col]

rows_original = []
rows_latin = []

print(f"Обрабатываем {N} строк (обрезка до {MAX_CHARS} символов + транслитерация)...")
idx = 1
for example in data:
    if len(rows_original) >= N:
        break
    text = str(example[text_col]).strip()
    if not text:
        continue
    text = truncate_text(text)
    row_meta = [example[c] for c in extra_cols]
    rows_original.append([idx, text] + row_meta)
    rows_latin.append([idx, transliterate(text)] + row_meta)
    if idx % 500 == 0:
        print(f"  {idx}/{N}...")
    idx += 1

print(f"Строк собрано: {len(rows_original)}")

print("\nПримеры транслитерации:")
for i in range(min(3, len(rows_original))):
    orig = rows_original[i][1][:80]
    lat = rows_latin[i][1][:80]
    print(f"  {orig}")
    print(f"  → {lat}\n")

# ── Создаём Workbook ──────────────────────────────────────────────────────
wb = Workbook()

col_w = {
    "A": 7,
    "B": 90,
    **{get_column_letter(i + 3): 15 for i in range(len(extra_cols))},
}

ws1 = wb.active
ws1.title = "Russian (Original)"
ws1.append(["#", text_col] + extra_cols)
write_rows(ws1, rows_original)
style_sheet(ws1, col_w)

ws2 = wb.create_sheet("Russian (Latin)")
ws2.append(["#", f"{text_col}_latin"] + extra_cols)
write_rows(ws2, rows_latin)
style_sheet(ws2, col_w)

# ── Сохраняем ─────────────────────────────────────────────────────────────
out_path = PROJECT_ROOT / "data" / "russian_latin_dataset.xlsx"
tmp_path = out_path.with_suffix(".tmp.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(tmp_path)
if out_path.exists():
    out_path.unlink()
tmp_path.rename(out_path)
print(f"Готово! Файл сохранён: {out_path}")
print(f"  Лист 1 «Russian (Original)»  — {len(rows_original)} строк")
print(f"  Лист 2 «Russian (Latin)»     — {len(rows_latin)} строк")
