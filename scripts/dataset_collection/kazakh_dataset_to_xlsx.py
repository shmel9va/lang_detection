"""
Скачивает казахский датасет KazParC с HuggingFace, берёт 2500 строк на казахском,
транслитерирует кириллицу → латиница по стандарту Республики Казахстан (апрель 2021).

Датасет: issai/kazparc (subset: kazparc_raw)
Колонки: id, kk, en, ru, tr, domain

Для доступа к датасету нужно принять условия на HuggingFace:
    https://huggingface.co/datasets/issai/kazparc

Запуск:
    py scripts/dataset_collection/kazakh_dataset_to_xlsx.py

Результат: data/kazakh_dataset.xlsx
  Лист 1 — Kazakh (Original):         2500 строк на кириллице
  Лист 2 — Kazakh (Transliterated Latin): 2500 строк на латинице
"""

from pathlib import Path
from datasets import load_dataset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]

# ── Транслитерация казахского (кириллица → латиница) ──────────────────────────
# Стандарт Республики Казахстан, апрель 2021 (президентский указ №637).
# 31 буква латинского алфавита + диграфы для русских заимствований.
# Источник: https://en.wikipedia.org/wiki/Kazakh_alphabets
#
# Ключевые особенности:
#   И, Й → İ/i  (как в турецком: İ — заглавная, i — строчная)
#   І    → I/ı   (I — заглавная, ı — строчная без точки)
#   Х, Һ → H/h  (слияние)
#   Ш    → Ş/ş   (седиль, не диграф sh)
TRANSLIT_MAP = {
    'а': 'a',   'ә': 'ä',   'б': 'b',   'в': 'v',   'г': 'g',
    'ғ': 'ğ',   'д': 'd',   'е': 'e',   'ё': 'io',  'ж': 'j',
    'з': 'z',   'и': 'i',   'й': 'i',   'к': 'k',   'қ': 'q',
    'л': 'l',   'м': 'm',   'н': 'n',   'ң': 'ñ',   'о': 'o',
    'ө': 'ö',   'п': 'p',   'р': 'r',   'с': 's',   'т': 't',
    'у': 'u',   'ұ': 'ū',   'ү': 'ü',   'ф': 'f',   'х': 'h',
    'һ': 'h',   'ц': 'ts',  'ч': 'ch',  'ш': 'ş',   'щ': 'şş',
    'ъ': '',    'ы': 'y',   'і': 'ı',   'ь': '',    'э': 'e',
    'ю': 'iu',  'я': 'ia',

    'А': 'A',   'Ә': 'Ä',   'Б': 'B',   'В': 'V',   'Г': 'G',
    'Ғ': 'Ğ',   'Д': 'D',   'Е': 'E',   'Ё': 'İo',  'Ж': 'J',
    'З': 'Z',   'И': 'İ',   'Й': 'İ',   'К': 'K',   'Қ': 'Q',
    'Л': 'L',   'М': 'M',   'Н': 'N',   'Ң': 'Ñ',   'О': 'O',
    'Ө': 'Ö',   'П': 'P',   'Р': 'R',   'С': 'S',   'Т': 'T',
    'У': 'U',   'Ұ': 'Ū',   'Ү': 'Ü',   'Ф': 'F',   'Х': 'H',
    'Һ': 'H',   'Ц': 'Ts',  'Ч': 'Ch',  'Ш': 'Ş',   'Щ': 'Şş',
    'Ъ': '',    'Ы': 'Y',   'І': 'I',   'Ь': '',    'Э': 'E',
    'Ю': 'İu',  'Я': 'İa',
}


def transliterate(text: str) -> str:
    return ''.join(TRANSLIT_MAP.get(ch, ch) for ch in text)


# ── Стиль ─────────────────────────────────────────────────────────────────────
HEADER_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL  = PatternFill('solid', start_color='0D47A1')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
ROW_FONT     = Font(name='Arial', size=10)
ROW_ALIGN    = Alignment(wrap_text=True, vertical='top')


def style_sheet(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def write_rows(ws, rows):
    for r_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 30
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = ROW_FONT
            cell.alignment = ROW_ALIGN


# ── Загрузка датасета ─────────────────────────────────────────────────────────
print("Загружаем датасет issai/kazparc (subset: kazparc_raw, streaming) ...")
ds = load_dataset("issai/kazparc", "kazparc_raw", streaming=True, trust_remote_code=True)

split = list(ds.keys())[0]
data  = ds[split]

col_names = data.column_names
print(f"Колонки: {col_names}")

text_col = 'kk'
extra_cols = [c for c in col_names if c != text_col]

N = 2500
rows_original       = []
rows_transliterated = []

print(f"Обрабатываем {N} строк ...")
idx = 1
for example in data:
    if len(rows_original) >= N:
        break
    text = str(example[text_col]).strip()
    if not text:
        continue
    row_meta = [example[c] for c in extra_cols]
    rows_original.append([idx, text] + row_meta)
    rows_transliterated.append([idx, transliterate(text)] + row_meta)
    if idx % 200 == 0:
        print(f"  {idx}/{N}...")
    idx += 1

print(f"Строк собрано: {len(rows_original)}")

# ── Создаём Workbook ──────────────────────────────────────────────────────────
wb = Workbook()

col_w = {
    'A': 7,
    'B': 90,
    **{get_column_letter(i + 3): 15 for i in range(len(extra_cols))},
}

ws1 = wb.active
ws1.title = "Kazakh (Original)"
ws1.append(['#', text_col] + extra_cols)
write_rows(ws1, rows_original)
style_sheet(ws1, col_w)

ws2 = wb.create_sheet("Kazakh (Transliterated Latin)")
ws2.append(['#', f'{text_col}_latin'] + extra_cols)
write_rows(ws2, rows_transliterated)
style_sheet(ws2, col_w)

# ── Сохраняем ─────────────────────────────────────────────────────────────────
out_path = PROJECT_ROOT / "data" / "kazakh_dataset.xlsx"
tmp_path = out_path.with_suffix(".tmp.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(tmp_path)
if out_path.exists():
    out_path.unlink()
tmp_path.rename(out_path)
print(f"Готово! Файл сохранён: {out_path}")
print(f"  Лист 1 «Kazakh (Original)»           — {len(rows_original)} строк")
print(f"  Лист 2 «Kazakh (Transliterated Latin)» — {len(rows_transliterated)} строк")
